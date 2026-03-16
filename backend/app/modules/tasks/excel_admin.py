from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from io import BytesIO
from typing import Any
from uuid import uuid4

from fastapi import Header, HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.modules.auth.models import User
from app.modules.tasks.models import Task, TaskAssignee, TaskVerifier


@dataclass(frozen=True)
class ExcelColumn:
    key: str
    label: str
    required: bool = False


EXCEL_COLUMNS: list[ExcelColumn] = [
    ExcelColumn("id", "ID задачи"),
    ExcelColumn("title", "Название *", required=True),
    ExcelColumn("description", "Описание"),
    ExcelColumn("creator_id", "ID создателя *", required=True),
    ExcelColumn("assignee_ids", "ID исполнителей"),
    ExcelColumn("verifier_ids", "ID проверяющих"),
    ExcelColumn("status", "Статус"),
    ExcelColumn("priority", "Приоритет"),
    ExcelColumn("due_date", "Дата выполнения *", required=True),
    ExcelColumn("due_time", "Время выполнения"),
    ExcelColumn("completed_at", "Дата и время завершения"),
    ExcelColumn("is_recurring", "Повторяющаяся задача"),
    ExcelColumn("recurrence_type", "Тип повторения"),
    ExcelColumn("recurrence_interval", "Интервал повторения"),
    ExcelColumn("recurrence_days_of_week", "Дни недели"),
    ExcelColumn("recurrence_end_date", "Дата окончания повторения"),
    ExcelColumn("recurrence_master_task_id", "ID мастер-задачи повторения"),
    ExcelColumn("source_module", "Модуль-источник"),
    ExcelColumn("source_entity", "Сущность-источник"),
    ExcelColumn("source_url", "Ссылка-источник"),
    ExcelColumn("created_at", "Создано"),
    ExcelColumn("updated_at", "Обновлено"),
]

LABEL_TO_KEY = {col.label: col.key for col in EXCEL_COLUMNS}
HEADER_ALIASES = {
    "created_by_user_id": "creator_id",
    "assignee_user_ids": "assignee_ids",
    "verifier_user_ids": "verifier_ids",
}
REQUIRED_KEYS = {c.key for c in EXCEL_COLUMNS if c.required}


@dataclass
class ImportErrorItem:
    row: int
    message: str


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[ImportErrorItem] | None = None

    def to_dict(self, total_rows: int = 0) -> dict[str, Any]:
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": [{"row": item.row, "message": item.message} for item in (self.errors or [])],
            "total_rows": total_rows,
        }


def validate_admin_pin(x_tasks_admin_pin: str | None = Header(default=None)) -> None:
    expected = os.getenv("TASKS_ADMIN_PIN", "0000")
    if (x_tasks_admin_pin or "") != expected:
        raise HTTPException(status_code=403, detail="Invalid admin PIN")


def _as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _parse_date_value(value: Any, field: str) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _as_text(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"{field}: ожидается формат YYYY-MM-DD") from exc


def _parse_time_value(value: Any, field: str) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    text = _as_text(value)
    if not text:
        return None
    try:
        return time.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"{field}: ожидается формат HH:MM") from exc


def _parse_datetime_value(value: Any, field: str) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(second=0, microsecond=0, tzinfo=timezone.utc)
    text = _as_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
    except ValueError as exc:
        raise ValueError(f"{field}: ожидается формат YYYY-MM-DD HH:MM") from exc


def _parse_bool(value: Any, default: bool = False) -> bool:
    text = _as_text(value).lower()
    if not text:
        return default
    if text in {"true", "1", "yes", "да"}:
        return True
    if text in {"false", "0", "no", "нет"}:
        return False
    raise ValueError("Повторяющаяся задача: допустимы TRUE/FALSE")


def _parse_int(value: Any, field: str, default: int | None = None) -> int | None:
    text = _as_text(value)
    if not text:
        return default
    try:
        return int(text)
    except ValueError as exc:
        raise ValueError(f"{field}: ожидается число") from exc


def _parse_user_ids(value: Any, field: str) -> list[int]:
    text = _as_text(value)
    if not text:
        return []
    result: list[int] = []
    for chunk in text.split(","):
        item = chunk.strip()
        if not item:
            continue
        try:
            result.append(int(item))
        except ValueError as exc:
            raise ValueError(f"{field}: неверный ID '{item}'") from exc
    return sorted(set(result))


def _validate_users_exist(db: Session, user_ids: list[int], field: str, cache: dict[int, bool]) -> None:
    for user_id in user_ids:
        exists = cache.get(user_id)
        if exists is None:
            exists = db.scalar(select(User.id).where(User.id == user_id).limit(1)) is not None
            cache[user_id] = exists
        if not exists:
            raise ValueError(f"{field}: пользователь {user_id} не найден")


def _parse_status(value: Any) -> str:
    status = _as_text(value).lower() or "active"
    if status == "done_pending_verify":
        raise ValueError("Статус done_pending_verify импортировать нельзя")
    if status not in {"active", "done"}:
        raise ValueError("Статус: допустимы active/done")
    return status


def _parse_priority(value: Any) -> str:
    priority = _as_text(value).lower() or "normal"
    if priority in {"urgent", "normal", "very_urgent"}:
        return {"urgent": "high", "normal": "normal", "very_urgent": "high"}[priority]
    if priority not in {"low", "normal", "high"}:
        raise ValueError("Приоритет: допустимы low/normal/high")
    return priority


def _priority_to_db(value: str) -> str:
    return {"low": "normal", "normal": "normal", "high": "very_urgent"}.get(value, "normal")


def _normalize_headers(raw_headers: list[Any]) -> list[str]:
    headers: list[str] = []
    for value in raw_headers:
        text = _as_text(value)
        key = LABEL_TO_KEY.get(text, text)
        key = HEADER_ALIASES.get(key, key)
        headers.append(key)
    return headers


def _rows_from_sheet(sheet) -> tuple[list[str], list[dict[str, Any]]]:
    raw_headers = [cell.value for cell in sheet[1]]
    headers = _normalize_headers(raw_headers)
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_map = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers)) if headers[idx]}
        row_map["_row_number"] = row_number
        rows.append(row_map)
    return headers, rows


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    ws.append([c.label for c in EXCEL_COLUMNS])
    for col_idx, col in enumerate(EXCEL_COLUMNS, start=1):
        if col.required:
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

    ws.append([
        "",
        "Проверить документы",
        "Пример заполненной задачи",
        "1",
        "1,2",
        "",
        "active",
        "normal",
        "2026-03-10",
        "10:30",
        "",
        "FALSE",
        "",
        "1",
        "",
        "",
        "",
        "counterparties",
        "contract",
        "https://example.local/task/1",
        "",
        "",
    ])

    status_validation = DataValidation(type="list", formula1='"active,done"', allow_blank=True)
    priority_validation = DataValidation(type="list", formula1='"low,normal,high"', allow_blank=True)
    due_date_validation = DataValidation(type="date", allow_blank=False)
    due_time_validation = DataValidation(type="time", allow_blank=True)
    completed_at_validation = DataValidation(type="date", allow_blank=True)
    recurrence_interval_validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True)
    creator_validation = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=False)

    status_validation.error = "Допустимы только active или done"
    priority_validation.error = "Допустимы только low, normal, high"
    due_date_validation.error = "Введите дату выполнения в формате даты Excel"
    creator_validation.error = "Введите числовой ID сотрудника"
    recurrence_interval_validation.error = "Введите целое число >= 1"

    ws.add_data_validation(status_validation)
    ws.add_data_validation(priority_validation)
    ws.add_data_validation(due_date_validation)
    ws.add_data_validation(due_time_validation)
    ws.add_data_validation(completed_at_validation)
    ws.add_data_validation(recurrence_interval_validation)
    ws.add_data_validation(creator_validation)

    status_validation.add("G2:G1048576")
    priority_validation.add("H2:H1048576")
    due_date_validation.add("I2:I1048576")
    due_time_validation.add("J2:J1048576")
    completed_at_validation.add("K2:K1048576")
    recurrence_interval_validation.add("N2:N1048576")
    creator_validation.add("D2:D1048576")

    ws["D1"].comment = Comment("Введите ID сотрудника из системы", "Codex")
    ws["I1"].comment = Comment("Дата, на которую задача должна быть выполнена", "Codex")
    ws["E1"].comment = Comment("ID исполнителей через запятую", "Codex")
    ws["F1"].comment = Comment("ID проверяющих через запятую", "Codex")

    note = wb.create_sheet("Инструкция")
    note.append(["Как пользоваться шаблоном"])
    notes = [
        "1) Заполните строки ниже заголовка. Обязательные поля выделены жирным и отмечены *.",
        "2) ID исполнителей/проверяющих и дни недели указывайте через запятую.",
        "3) Статус допускается только active или done. done_pending_verify импортировать нельзя.",
        "4) Для status=done поле 'Дата и время завершения' обязательно.",
        "5) Если ID задачи указан и найден — задача обновится. Если пустой — создастся новая.",
    ]
    for item in notes:
        note.append([item])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_tasks_workbook(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Экспорт задач"
    ws.append([c.label for c in EXCEL_COLUMNS])

    tasks = db.scalars(select(Task).order_by(Task.created_at.asc(), Task.id.asc())).all()
    task_ids = [t.id for t in tasks]
    assignee_rows = db.execute(select(TaskAssignee.task_id, TaskAssignee.user_id).where(TaskAssignee.task_id.in_(task_ids))).all() if task_ids else []
    verifier_rows = db.execute(select(TaskVerifier.task_id, TaskVerifier.user_id).where(TaskVerifier.task_id.in_(task_ids))).all() if task_ids else []
    assignee_map: dict[str, list[int]] = {}
    verifier_map: dict[str, list[int]] = {}
    for task_id, user_id in assignee_rows:
        assignee_map.setdefault(task_id, []).append(user_id)
    for task_id, user_id in verifier_rows:
        verifier_map.setdefault(task_id, []).append(user_id)

    for task in tasks:
        ws.append([
            task.id,
            task.title,
            task.description or "",
            task.created_by_user_id,
            ",".join(str(v) for v in sorted(set(assignee_map.get(task.id, [])))),
            ",".join(str(v) for v in sorted(set(verifier_map.get(task.id, [])))),
            task.status,
            task.priority or "",
            task.due_date.isoformat() if task.due_date else "",
            task.due_time.strftime("%H:%M") if task.due_time else "",
            task.completed_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M") if task.completed_at else "",
            "TRUE" if task.is_recurring else "FALSE",
            task.recurrence_type or "",
            task.recurrence_interval or "",
            task.recurrence_days_of_week or "",
            task.recurrence_end_date.isoformat() if task.recurrence_end_date else "",
            task.recurrence_master_task_id or "",
            task.source_module or "",
            task.source_type or "",
            task.source_id or "",
            task.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M"),
            "",
        ])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _validate_row(db: Session, row: dict[str, Any], user_cache: dict[int, bool]) -> tuple[dict[str, Any], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    normalized: dict[str, Any] = {k: _as_text(v) for k, v in row.items() if not k.startswith("_")}
    try:
        title = _as_text(row.get("title"))
        if not title:
            raise ValueError("Название: обязательное поле")
        creator_id = _parse_int(row.get("creator_id"), "ID создателя")
        if creator_id is None:
            raise ValueError("ID создателя: обязательное поле")
        _validate_users_exist(db, [creator_id], "ID создателя", user_cache)

        assignee_ids = _parse_user_ids(row.get("assignee_ids"), "ID исполнителей") or [creator_id]
        verifier_ids = _parse_user_ids(row.get("verifier_ids"), "ID проверяющих")
        _validate_users_exist(db, assignee_ids, "ID исполнителей", user_cache)
        _validate_users_exist(db, verifier_ids, "ID проверяющих", user_cache)

        status = _parse_status(row.get("status"))
        completed_at = _parse_datetime_value(row.get("completed_at"), "Дата и время завершения")
        if status == "done" and completed_at is None:
            raise ValueError("Для статуса done заполните 'Дата и время завершения'")

        priority = _parse_priority(row.get("priority"))
        due_date = _parse_date_value(row.get("due_date"), "Дата выполнения")
        if due_date is None:
            raise ValueError("не указана дата выполнения задачи")
        due_time = _parse_time_value(row.get("due_time"), "Время выполнения")
        is_recurring = _parse_bool(row.get("is_recurring"), False)
        recurrence_interval = _parse_int(row.get("recurrence_interval"), "Интервал повторения", 1)
        if recurrence_interval is not None and recurrence_interval < 1:
            raise ValueError("Интервал повторения должен быть >= 1")
        recurrence_end_date = _parse_date_value(row.get("recurrence_end_date"), "Дата окончания повторения")

        task_id = _as_text(row.get("id"))
        existing = bool(task_id and db.scalar(select(Task.id).where(Task.id == task_id).limit(1)))

        normalized.update(
            {
                "id": task_id,
                "title": title,
                "creator_id": creator_id,
                "assignee_ids": ",".join(str(i) for i in assignee_ids),
                "verifier_ids": ",".join(str(i) for i in verifier_ids),
                "status": status,
                "priority": priority,
                "due_date": due_date.isoformat() if due_date else "",
                "due_time": due_time.strftime("%H:%M") if due_time else "",
                "completed_at": completed_at.strftime("%Y-%m-%d %H:%M") if completed_at else "",
                "is_recurring": "TRUE" if is_recurring else "FALSE",
                "recurrence_interval": str(recurrence_interval or 1),
                "recurrence_end_date": recurrence_end_date.isoformat() if recurrence_end_date else "",
                "_import_action": "update" if existing else "create",
            }
        )
    except ValueError as exc:
        errors.append(str(exc))
    return normalized, errors, warnings


def build_import_preview(db: Session, upload: UploadFile) -> dict[str, Any]:
    if not upload.filename or not upload.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    sheet = load_workbook(filename=BytesIO(upload.file.read()), data_only=True).active
    _, rows = _rows_from_sheet(sheet)
    user_cache: dict[int, bool] = {}
    preview_rows = []
    invalid_rows = 0
    for row in rows:
        normalized, errors, warnings = _validate_row(db, row, user_cache)
        if all(_as_text(v) == "" for k, v in normalized.items() if not k.startswith("_")):
            continue
        if errors:
            invalid_rows += 1
        preview_rows.append(
            {
                "row_number": row["_row_number"],
                "values": normalized,
                "errors": errors,
                "warnings": warnings,
            }
        )
    total = len(preview_rows)
    return {
        "columns": [
            {
                "key": c.key,
                "label": c.label,
                "required": c.required,
                "field_type": (
                    "date" if c.key == "due_date" else
                    "time" if c.key == "due_time" else
                    "datetime" if c.key == "completed_at" else
                    "number" if c.key in {"creator_id", "recurrence_interval"} else
                    "select" if c.key in {"status", "priority", "creator_id", "assignee_ids", "verifier_ids"} else
                    "text"
                ),
                "searchable": c.key in {"creator_id", "assignee_ids", "verifier_ids"},
                "options": (
                    [{"value": "active", "label": "active"}, {"value": "done", "label": "done"}] if c.key == "status" else
                    [{"value": "low", "label": "low"}, {"value": "normal", "label": "normal"}, {"value": "high", "label": "high"}] if c.key == "priority" else
                    []
                ),
            }
            for c in EXCEL_COLUMNS
        ],
        "users": [{"id": user.id, "username": user.username} for user in db.scalars(select(User).order_by(User.username.asc())).all()],
        "rows": preview_rows,
        "row_errors": [{"row": r["row_number"], "errors": r["errors"]} for r in preview_rows if r["errors"]],
        "row_warnings": [{"row": r["row_number"], "warnings": r["warnings"]} for r in preview_rows if r["warnings"]],
        "total_rows": total,
        "create_rows": sum(1 for row in preview_rows if row["values"].get("_import_action") == "create"),
        "update_rows": sum(1 for row in preview_rows if row["values"].get("_import_action") == "update"),
        "valid_rows": total - invalid_rows,
        "invalid_rows": invalid_rows,
    }


def import_tasks_from_preview(db: Session, rows: list[dict[str, Any]]) -> dict[str, Any]:
    result = ImportResult(errors=[])
    total_rows = len(rows)
    user_cache: dict[int, bool] = {}
    for index, item in enumerate(rows, start=1):
        row = item.get("values", item)
        try:
            normalized, errors, _ = _validate_row(db, row, user_cache)
            if errors:
                raise ValueError("; ".join(errors))
            creator_id = int(normalized["creator_id"])
            assignee_ids = _parse_user_ids(normalized.get("assignee_ids"), "ID исполнителей") or [creator_id]
            verifier_ids = _parse_user_ids(normalized.get("verifier_ids"), "ID проверяющих")
            status = normalized["status"]
            task_id = normalized.get("id")
            existing = db.scalar(select(Task).where(Task.id == task_id).limit(1)) if task_id else None

            payload = {
                "title": normalized["title"],
                "description": _as_text(normalized.get("description")) or None,
                "due_date": _parse_date_value(normalized.get("due_date"), "Дата выполнения"),
                "due_time": _parse_time_value(normalized.get("due_time"), "Время выполнения"),
                "priority": _priority_to_db(normalized["priority"]),
                "status": status,
                "created_by_user_id": creator_id,
                "completed_at": _parse_datetime_value(normalized.get("completed_at"), "Дата и время завершения") if status == "done" else None,
                "is_recurring": _parse_bool(normalized.get("is_recurring"), False),
                "recurrence_type": _as_text(normalized.get("recurrence_type")) or None,
                "recurrence_interval": _parse_int(normalized.get("recurrence_interval"), "Интервал повторения", 1),
                "recurrence_days_of_week": _as_text(normalized.get("recurrence_days_of_week")) or None,
                "recurrence_end_date": _parse_date_value(normalized.get("recurrence_end_date"), "Дата окончания повторения"),
                "recurrence_master_task_id": _as_text(normalized.get("recurrence_master_task_id")) or None,
                "source_module": _as_text(normalized.get("source_module")) or None,
                "source_type": _as_text(normalized.get("source_entity")) or None,
                "source_id": _as_text(normalized.get("source_url")) or None,
            }

            if existing:
                for key, value in payload.items():
                    setattr(existing, key, value)
                db.query(TaskAssignee).filter(TaskAssignee.task_id == existing.id).delete()
                db.query(TaskVerifier).filter(TaskVerifier.task_id == existing.id).delete()
                for uid in assignee_ids:
                    db.add(TaskAssignee(task_id=existing.id, user_id=uid))
                for uid in verifier_ids:
                    db.add(TaskVerifier(task_id=existing.id, user_id=uid))
                result.updated += 1
            else:
                new_task = Task(
                    id=task_id or str(uuid4()),
                    created_at=datetime.now(timezone.utc),
                    verified_at=None,
                    source_counterparty_id=None,
                    source_trigger_id=None,
                    recurrence_state="active",
                    is_hidden=False,
                    **payload,
                )
                db.add(new_task)
                db.flush()
                for uid in assignee_ids:
                    db.add(TaskAssignee(task_id=new_task.id, user_id=uid))
                for uid in verifier_ids:
                    db.add(TaskVerifier(task_id=new_task.id, user_id=uid))
                result.created += 1
        except Exception as exc:
            result.errors.append(ImportErrorItem(row=index, message=str(exc)))
            result.skipped += 1

    db.commit()
    return result.to_dict(total_rows=total_rows)
